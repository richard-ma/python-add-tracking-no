#!/usr/bin/env python
# encoding: utf-8

import os
import csv
import glob

from openpyxl import load_workbook
from shutil import copy
from pprint import pprint

class DataFile(object):

    def __init__(self, filename, ext_filter=None):
        self.filename = filename
        self.ext_filter = ext_filter
        self.filetype = self._getFileType()

    def _getFileType(self):
        ext = os.path.splitext(self.filename)[-1]

        # check filename
        if self.ext_filter is not None and ext not in self.ext_filter:
            raise Exception("本程序不支持扩展名为%s的文件%s" % (ext, self.filename))
        
        return ext
    
    def read(self):
        self.workbook = load_workbook(filename = self.filename)
        return self

    def write(self):    
        self.workbook.save(self.filename)
        return self


class ShipDataFile(DataFile):
    def get_data(self):
        ws = self.workbook.active
        d = dict()
        for row in ws.iter_rows(min_row=2): # remove title line
            order_id, tracking_no = str(row[0].value), row[1].value
            if '-' in order_id:
                order_id = order_id.split('-')[-1]
            d[order_id] = tracking_no
            
            print("Loading Relationship: %s -> %s" % (order_id, tracking_no))
        return d
        

class ExpressDataFile(DataFile):
    def add_tracking_no(self, tracking_no_data):
        notfound = list()
        used = set()
        processing_count, used_count, notfound_count = 0, 0, 0
        line_no = 1
        for row in self.workbook.active.iter_rows(min_row=2):
            line_no += 1
            
            order_id_value = row[0].value
            order_id = str(order_id_value)
            
            if order_id_value is None:
                continue
            
            
            # unmerge cell
            self.workbook.active.unmerge_cells('M%d:M%d' % (line_no, line_no+12))
            
            if order_id in tracking_no_data.keys():
                self.workbook.active['M%d' % (line_no)] = tracking_no_data[order_id]
                
                used.add(order_id)
                used_count += 1
            else:
                notfound.append(order_id)
                notfound_count += 1
            
            #merge cells
            self.workbook.active.merge_cells('M%d:M%d' % (line_no, line_no+12))
                        
            processing_count += 1
        
            print("[%dp/%du/%dn] order_id: %s, tracking_no: %s" %(
                processing_count,
                used_count,
                notfound_count,
                order_id,
                self.workbook.active['M%d' % (line_no)]
            ))
                
        unused = list()
        if len(used) < len(tracking_no_data):
            unused = set(tracking_no_data.keys()) - used
            unused = list(unused)
        
        return notfound, unused
                    

class RecordDataFile():
    def __init__(self, filename):
        self.filename = filename

    def write(self, data):
        with open(self.filename, 'w+', encoding='GBK') as f:
            for line in data:
                f.write(line + '\r\n')

if __name__ == "__main__":
    
    record_filename = 'record.txt'
    
    ship_filename = 'test/20210808-订单导出.xlsx'
    express_filename = 'test/0808(1020016-804)_8535.xlsx'
    
    sdf = ShipDataFile(ship_filename, ['.xlsx']).read()
    ship_data = sdf.get_data()
    
    ef = ExpressDataFile(express_filename, ['.xlsx']).read()
    notfound, unused = ef.add_tracking_no(ship_data)
    ef.write()
    
    if len(notfound) > 0 or len(unused) > 0:
        data = [
            '没有快递单号的订单ID',
            *notfound,
            '快递单列表中没有使用的订单ID',
            *unused
        ]
        rf = RecordDataFile(record_filename)
        rf.write(data)
